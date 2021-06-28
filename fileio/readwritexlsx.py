"""
Created on 06/11/2020
@author: ingridtveten

"""

import openpyxl


# =====================================
#   BASIC XLSX FUNCTIONS
# =====================================

def read_workbook(filename):

    wb = openpyxl.load_workbook(filename=filename, read_only=True)
    ws = wb.worksheets[0]

    data = [[]]
    i = 0
    for row in ws.rows:
        for c in row:
            data[i].append(c.value)  # openpyxl uses 1-indexing

        i += 1
        if i != ws.max_row:
            data.append([])

    return data


def xlsx_data_to_dict(data):
    d = {}
    for idx, item in enumerate(data[0]):
        item = str(item)    # Read headers
        col = []
        for r in range(1, len(data)):  # For each row
            col.append(data[r][idx])
        d[item] = col  # Read data

    return d


# =====================================
#   COMPLICATION DATA FUNCTIONS
# =====================================

def check_complication_values(data_dict):

    for key, lst in data_dict.items():
        for idx, v in enumerate(lst):

            if v == '#NULL!':
                v = None
                lst[idx] = v

        data_dict[key] = lst

    return data_dict


def read_xlsx_complication_data(filename):

    complication_data = read_workbook(filename)
    complications = xlsx_data_to_dict(complication_data)
    complications = check_complication_values(complications)
    return complications


# =====================================
#   TEST FUNCTION
# =====================================

def test_xlsx_writer():
    from openpyxl import Workbook

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")



if __name__ == "__main__":
    from constants import QOL_FILE_PATH

    #test_xlsx_writer()
    complication_data = read_xlsx_complication_data(QOL_FILE_PATH)
